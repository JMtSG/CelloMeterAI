print("""\n
      ____________________________________
     |                                    |
     |         CelloMeterAI: v1.1         |
     |____________________________________|""")
print("      - Developed by John Mai · Apr 2023 -\n")
import warnings
warnings.filterwarnings("ignore")
import time
PROGRAM_START_TIME = time.time()

import sys
# wait for [enter] key from user before exiting
def exit_hang():
    input("\n\nPress [Enter] to exit")
def show_exception_and_exit(exc_type, exc_value, tb):	# run this on error
    import traceback
    print('\n\n')
    traceback.print_exception(exc_type, exc_value, tb)
    with open('ERROR.LOG', 'w') as f:
        traceback.print_exception(exc_type, exc_value, tb, file=f)
    exit_hang()
    sys.exit(-1)

sys.excepthook = show_exception_and_exit
print("Initialising... ", end='', flush=True)
import numpy as np
import cv2
import torch
import os
import glob
import copy
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import yaml
import concurrent.futures
from typing import List, Dict, Union

from image_processor import ImageProcessor
from CNN import CNN_Detector

print("Done.\n")

STATUS_BAR_LEN = 50

# Load settings
with open('settings.yaml', 'r') as stream:
    try:
        settings = yaml.safe_load(stream)
        COLOUR_REGULAR_CELL = tuple(settings['COLOURS']['REGULAR_CELL'][::-1])
        COLOUR_ABNORMAL_CELL = tuple(settings['COLOURS']['ABNORMAL_CELL'][::-1])
        COLOUR_BLOB_CLUSTER = tuple(settings['COLOURS']['BLOB_CLUSTER'][::-1])
    except yaml.YAMLError as e:
        print(e)

# Define whether we want to skip every 2nd, 3rd, etc. image when displaying cell detections
if settings['INFERENCE_DEVICE']=='cpu':
    # PROGRESS_IMG_DISP_STRIDE = settings['NUM_WORKERS']
    PROGRESS_IMG_DISP_STRIDE = 3
else:
    PROGRESS_IMG_DISP_STRIDE = 10

def get_IDs_to_discard(detections_tile_A, detections_tile_B):
    """
    Return a list of detection IDs from tile A that overlap with detections from tile B
    
    Args:
    - detections_tile_A (List[List]): List of detections in tile A, each detection is represented as [tile index, class enum, confidence score, bounding box coordinates]
    - detections_tile_B (List[List]): List of detections in tile B, each detection is represented as [tile index, class enum, confidence score, bounding box coordinates]
    
    Returns:
    - detection_IDs_to_discard_A (List): List of detection IDs to discard from tile A that overlap with detections from tile B
    """
    detection_IDs_to_discard_A = []
    for detection_A in detections_tile_A:
        for detection_B in detections_tile_B:
            percent_overlap = calc_rect_overlap(detection_A[4:8], detection_B[4:8])
            if percent_overlap>settings['MAX_OVERLAP_PERCENTAGE']:
                detection_IDs_to_discard_A.append(detection_A[1])
    return detection_IDs_to_discard_A

def calc_rect_overlap(coords_A, coords_B):
    """
    Calculate the percentage overlap between two rectangles
    
    Args:
    - coords_A (List): Coordinates of the first rectangle [x1, y1, x2, y2], where (x1,y1) represents the top left corner and (x2,y2) represents the bottom right corner.
    - coords_B (List): Coordinates of the second rectangle [x1, y1, x2, y2], where (x1,y1) represents the top left corner and (x2,y2) represents the bottom right corner.
    
    Returns:
    - overlap_percent (float): The percentage overlap between the two rectangles
    """
    x_left = max(coords_A[0], coords_B[0])
    y_top = max(coords_A[1], coords_B[1])
    x_right = min(coords_A[2], coords_B[2])
    y_bottom = min(coords_A[3], coords_B[3])
    
    # Calculate the width and height of the intersection rectangle
    intersection_width = max(0, x_right - x_left)
    intersection_height = max(0, y_bottom - y_top)
    
    # Calculate the area of the intersection rectangle
    intersection_area = intersection_width * intersection_height
    
    # Calculate the areas of both bounding boxes
    bb1_area = (coords_B[2] - coords_B[0]) * (coords_B[3] - coords_B[1])
    bb2_area = (coords_B[2] - coords_B[0]) * (coords_B[3] - coords_B[1])
    
    # Calculate the percentage overlap (overlap_area / total area)
    # overlap_percent = (intersection_area / float(bb1_area + bb2_area - intersection_area))
    # Calculate the percentage overlap (100% if any bounding box is enveloped by another)
    overlap_percent = max(intersection_area/float(bb1_area),
                          intersection_area/float(bb2_area))
    
    return overlap_percent


def generate_report(input_data, output_fname):
    """
    Generate a report of cell count data and save it to an Excel file
    
    Args:
    - input_data (Dict): Dictionary containing cell count data for each image, with image names as keys and a list of cell counts as values.
    - output_fname (str): File name (including path) to save the Excel report to.
    
    Returns:
    - None
    """
    data = copy.deepcopy(input_data)

    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define the header row
    header = ['File name', 'Normal Cells', 'Abnormal Cells', 'Cell Clusters', 'Total Count']

    # Set the background color and font style for the header row
    font_reg = Font(name='Arial', bold=False)
    font_bold = Font(name='Arial', bold=True)

    # Write the header row to the worksheet
    for col_num, value in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num, value=value)
        if col_num==2:  # normal cells
            cell.fill = PatternFill(start_color='FF66FF', end_color='FF66FF', fill_type='solid')
        elif col_num==3:  # abnormal cells
            cell.fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
        elif col_num==4:  # blob clusters
            cell.fill = PatternFill(start_color='FFB366', end_color='FFB366', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')    # grey
        cell.font = font_bold

    # Calculate the total number of cells in each image
    for key in data:
        data[key].append(sum(data[key][:3]))

    # Write the data to the worksheet
    for row_num, key in enumerate(data, 2):
        row_data = [key] + data[key]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.font = font_reg

    # Calculate the totals for each column
    normal_cells_total = sum([data[key][0] for key in data])
    abnormal_cells_total = sum([data[key][1] for key in data])
    blob_clusters_total = sum([data[key][2] for key in data])
    total_cells_total = sum([data[key][3] for key in data])

    # Add a total row at the bottom of the worksheet
    total_row = ['Final Totals', normal_cells_total, abnormal_cells_total, blob_clusters_total, total_cells_total]
    for col_num, value in enumerate(total_row, 1):
        cell = ws.cell(row=len(data)+2, column=col_num, value=value)
        cell.font = font_bold

    # Auto-adjust the column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Save the workbook
    wb.save(output_fname)

def run_inference(tile_idx, tile, tile_x, tile_y, cnn_detector):
    """
    Run the CNN detector on a given tile and return the detections and the tile coordinates.

    Args:
    - tile_idx (int): The index of the tile.
    - tile (numpy array): The tile to run the detector on.
    - tile_x (int): The x-coordinate of the tile in the image.
    - tile_y (int): The y-coordinate of the tile in the image.
    - cnn_detector: The CNN detector to use.

    Returns:
    - Tuple: The detections found by the CNN and the coordinates of the tile.
    """
    ### Run the CNN ###
    detections = cnn_detector.detect(tile)
    return detections, (tile_idx, tile_x, tile_y)

def display_detection_progression(img, detection, detection_id, title, bbox_coords):
    """
    Annotate the detections found on a tile on the full image (unscaled).

    Args:
    - img (numpy array): The full image to annotate the detections on.
    - detection (list): The detection to annotate.
    - detection_id (int): The ID of the detection.
    - title (str): The title of the image.
    - bbox_coords (list): The bounding box coordinates of the detection.

    Returns:
    - None
    """
    # Annotate the detections found on this tile on the full image (unscaled)
    # Need to scale the coordinates back to the original image size
    x1,y1,x2,y2 = np.round(bbox_coords/settings['ZOOM_FACTOR']).astype(int)
    if detection[0]==0:     # normal cell
        cv2.rectangle(img, (x1, y1), (x2, y2), COLOUR_REGULAR_CELL, 1)
    elif detection[0]==1:   # abnormal looking cell
        cv2.rectangle(img, (x1, y1), (x2, y2), COLOUR_ABNORMAL_CELL, 1)
    elif detection[0]==2:   # blob cluster of cells
        cv2.rectangle(img, (x1, y1), (x2, y2), COLOUR_BLOB_CLUSTER, 1)
    if detection_id%PROGRESS_IMG_DISP_STRIDE==0:   # update image every Nth detection
        cv2.imshow(title, img)
        cv2.waitKey(1)

def count_cells(input_img_fnames: List[str], image_processor: ImageProcessor,
                cnn_detector: CNN_Detector) -> Dict[str, Union[str, int]]:
    """
    Counts the number of cells present in a set of images. For each input image,
    the function loads the image, scales it, and generates tiles. For each tile,
    the CNN detector is run and the bounding boxes for detected cells are
    extracted. The function then consolidates these bounding boxes across all tiles,
    eliminates overlapping bounding boxes, and returns a summary of the number of
    cells found in the image.

    Args:
    - input_img_fnames (List[str]): A list of input image filenames.
    - image_processor (ImageProcessor): An instance of the ImageProcessor class
      for generating tiles from an image.
    - cnn_detector (CNNDetector): An instance of the CNNDetector class for running
      cell detection on an image tile.

    Returns:
    - cell_count_summary (Dict[str, Union[str, int]]): A dictionary containing the
      following keys and values:
      - 'filename': The name of the input image file.
      - 'num_cells': The total number of cells found in the image.
      - 'cell_positions': A list of tuples, where each tuple contains the x and y
        coordinates of a cell in the image.
    """

    # Load each image one at a time and process
    cell_count_summary = {}
    detection_id = 0    # unique serial ID for each detection across the whole image
    for file_idx,input_fname in enumerate(input_img_fnames):
        base_fname = os.path.basename(input_fname)
        print(f"Processing image ({file_idx+1}/{len(input_img_fnames)}): \"{base_fname}\"")
        # try:
        orig_img = cv2.imread(input_fname)
        display_img = orig_img.copy()
        orig_img_shape = orig_img.shape[:2]
        # Scale up the image
        img = cv2.resize(orig_img, (int(orig_img_shape[1] * settings['ZOOM_FACTOR']), 
                            int(orig_img_shape[0] * settings['ZOOM_FACTOR'])))
        img_tiles = image_processor.generate_tiles(img,
                                                    settings['CNN_INPUT_DIM'],
                                                    settings['TILE_OVERLAP_FACTOR'])
        
        cell_detections = []    # a list of tuples, (tile index, detection serial ID, class, confidence, x1, y1, x2, y2)
        if settings['NUM_WORKERS']>1:
            # Create a ThreadPoolExecutor with a maximum number of threads
            with concurrent.futures.ThreadPoolExecutor(max_workers=settings['NUM_WORKERS']) as executor:
                # Submit the processing of each img_tiles object to the executor and store the resulting future object
                futures = [executor.submit(run_inference, tile_idx, tile, tile_x, tile_y, cnn_detector) 
                        for tile_idx,(tile,(tile_x,tile_y), _) in enumerate(img_tiles)]
                # Wait for all futures to complete
                # concurrent.futures.wait(futures)      # Blocking delay while we wait for the tiles to be processed
                completed_tiles_count = 0
                while futures:                          # Delay that allows to run code in the main thread while we wait
                    # Check for tiles processed
                    completed_tiles, _ = concurrent.futures.wait(futures, timeout=0.001)
                    # Remove the completed futures from the list
                    futures = [f for f in futures if not f.done()]
                    completed_tiles_count += len(completed_tiles)
                    # Generate status bar and percentage indicator for displaying progress
                    # In this loop, maximum progress caps out at 90%
                    completion_pc = (completed_tiles_count)/len(img_tiles) * 0.9
                    hash_symbol_cnt = round(completion_pc*STATUS_BAR_LEN)
                    status_bar = '[' + '#'*hash_symbol_cnt + '-'*(STATUS_BAR_LEN-hash_symbol_cnt) + ']'
                    print(f"{status_bar} %d%%"%round(completion_pc*100.0), end='\r')
                    
                    # Loop through all the results if there are any
                    if completed_tiles:
                        for model_output in completed_tiles:
                            detections, (tile_idx, tile_x, tile_y) = model_output.result()
                            for detection in detections:
                                # Transform bounding box coordinates from tile frame to full image frame
                                x1 = int(detection[2]) + tile_x
                                y1 = int(detection[3]) + tile_y
                                x2 = int(detection[4]) + tile_x
                                y2 = int(detection[5]) + tile_y
                                # (tile index, serial ID, class enum, confidence score, bounding box coordinates)
                                cell_detections.append((tile_idx, detection_id, detection[0], detection[1], x1, y1, x2, y2))
                                detection_id += 1
                                # Annotate the detections found on this tile on the full image (unscaled)
                                if settings['LIVE_DISPLAY']:
                                    display_detection_progression(display_img, detection, detection_id, base_fname, np.r_[x1,y1,x2,y2])
        else:   # Single worker/thread
            for tile_idx,(tile,(tile_x,tile_y), adjacent_tile_idxs) in enumerate(img_tiles):
                # Generate status bar and percentage indicator for displaying progress
                # In this loop, maximum progress caps out at 90%
                completion_pc = (tile_idx+1)/len(img_tiles) * 0.9
                hash_symbol_cnt = round(completion_pc*STATUS_BAR_LEN)
                status_bar = '[' + '#'*hash_symbol_cnt + '-'*(STATUS_BAR_LEN-hash_symbol_cnt) + ']'
                print(f"{status_bar} %d%%"%round(completion_pc*100.0), end='\r')
                
                ### Run the CNN ###
                detections = cnn_detector.detect(tile)
                # Go through the detected objects
                for detection in detections:
                    # Transform bounding box coordinates from tile frame to full image frame
                    x1 = int(detection[2]) + tile_x
                    y1 = int(detection[3]) + tile_y
                    x2 = int(detection[4]) + tile_x
                    y2 = int(detection[5]) + tile_y
                    # (tile index, serial ID, class enum, confidence score, bounding box coordinates)
                    cell_detections.append((tile_idx, detection_id, detection[0], detection[1], x1, y1, x2, y2))
                    detection_id += 1
                    # Annotate the detections found on this tile on the full image (unscaled)
                    if settings['LIVE_DISPLAY']:
                        display_detection_progression(display_img, detection, detection_id, base_fname, np.r_[x1,y1,x2,y2])
        cell_detections = np.array(cell_detections)

        # Now eliminate cells that were counted twice; this happens when tiles overlap and the same cell shows up on more than one tile
        if cell_detections.shape[0]>0:
            # Create an array that contains entries of bounding box coordinates, the tile index, and
            # For each tile, check if any detection is too close to a detection in any tile to the right or below the tile
            detection_IDs_to_discard = []
            for tile_idx,(tile,(tile_x,tile_y), adjacent_tile_idxs) in enumerate(img_tiles):
                for adj_tile_idx in adjacent_tile_idxs:
                    # Complete the rest of the progress bar here
                    completion_pc = ((tile_idx+1)/len(img_tiles) * 0.1) + 0.9
                    hash_symbol_cnt = round(completion_pc*STATUS_BAR_LEN)
                    status_bar = '[' + '#'*hash_symbol_cnt + '-'*(STATUS_BAR_LEN-hash_symbol_cnt) + ']'
                    print(f"{status_bar} %d%%"%round(completion_pc*100.0), end='\r')
                    
                    # Discard the tuple of adjacent tile indexes
                    curr_tile_detections = cell_detections[cell_detections[:,0]==tile_idx]
                    adj_tile_detections = cell_detections[cell_detections[:,0]==adj_tile_idx]
                    # We now have two arrays of detections, where some detections may overlap with another
                    # Take the detections from the current tile, and discard any that overlap with the adjacent tile
                    discard_IDs = get_IDs_to_discard(curr_tile_detections, adj_tile_detections)
                    detection_IDs_to_discard.extend(discard_IDs)
                # Redraw the image with annotations removed from discarded detections
                if settings['LIVE_DISPLAY']:
                    mask = np.isin(cell_detections[:, 1].astype('int'), np.array(detection_IDs_to_discard).astype('int'))
                    unique_cell_detections = np.delete(cell_detections.copy(), np.where(mask), axis=0)
                    display_img = orig_img.copy()
                    for detection in unique_cell_detections:
                        # Need to scale the coordinates back to the original image size
                        x1, y1, x2, y2 = (detection[4:8]/settings['ZOOM_FACTOR']).astype('int')
                        if detection[2]==0:     # normal cell
                            cv2.rectangle(display_img, (x1, y1), (x2, y2), COLOUR_REGULAR_CELL, 1)
                        elif detection[2]==1:   # abnormal looking cell
                            cv2.rectangle(display_img, (x1, y1), (x2, y2), COLOUR_ABNORMAL_CELL, 1)
                        elif detection[2]==2:   # blob cluster of cells
                            cv2.rectangle(display_img, (x1, y1), (x2, y2), COLOUR_BLOB_CLUSTER, 1)
                    cv2.imshow(base_fname, display_img)
                    cv2.waitKey(1)
            ## Discard repeated cell detections
            # Create a boolean mask for rows to delete
            mask = np.isin(cell_detections[:, 1].astype('int'), np.array(detection_IDs_to_discard).astype('int'))
            unique_cell_detections = np.delete(cell_detections, np.where(mask), axis=0)
            # print(f"{len(unique_cell_detections)} unique_cell_detections")

            # Annotate main image with the cell detections, and count the total numbers of cells
            cell_count = {'normal': 0, 'abnormal': 0, 'cluster': 0}
            for detection in unique_cell_detections:
                x1, y1, x2, y2 = detection[4:8].astype('int')
                if detection[2]==0:     # normal cell
                    cell_count['normal'] += 1
                    cv2.rectangle(img, (x1, y1), (x2, y2), COLOUR_REGULAR_CELL, settings['ANNOTATION_THICKNESS'])
                elif detection[2]==1:   # abnormal looking cell
                    cell_count['abnormal'] += 1
                    cv2.rectangle(img, (x1, y1), (x2, y2), COLOUR_ABNORMAL_CELL, settings['ANNOTATION_THICKNESS'])
                elif detection[2]==2:   # blob cluster of cells
                    cell_count['cluster'] += 1
                    cv2.rectangle(img, (x1, y1), (x2, y2), COLOUR_BLOB_CLUSTER, settings['ANNOTATION_THICKNESS'])
            print(f"\nCell count: {cell_count['normal']} normal, {cell_count['abnormal']} abnormal, {cell_count['cluster']} clusters")
            cell_count_summary[f"{base_fname}"] = [cell_count['normal'], cell_count['abnormal'], cell_count['cluster']]
        else:
            print(f"[{'#'*STATUS_BAR_LEN}] 100%", end='\r')
            cell_count_summary[f"{base_fname}"] = [0, 0, 0]
            print(f"\nNo cells detected")

        # Output the an image regardless of whether there are detections
        output_img = cv2.resize(img, (settings['OUTPUT_IMG_W'],
                                        int(settings['OUTPUT_IMG_W']/img.shape[1] * img.shape[0])))
        cv2.imwrite(f"{settings['OUTPUT_FOLDER']}/{base_fname}", output_img)
        cv2.destroyAllWindows()
        # except RuntimeError as e:
        #     if "at non-singleton dimension 3" in str(e):
        #         print("*** ERROR: Possibly due to exceeding computational resources ***")
        #         print("Reverting to single-worker processing")
        #     else:
        #         print(f"RuntimeError: {e}")
        # except:
        #     print(f"Error on image {base_fname}")
        print('')
        
        # Generate an excel spreadsheet report for the data we have so far
        generate_report(cell_count_summary, f"{settings['OUTPUT_FOLDER']}/{settings['REPORT_FILE_NAME']}.xlsx")



if __name__=='__main__':
    image_processor = ImageProcessor()  # For tiling & un-tiling images
    cnn_detector = CNN_Detector(f"./{settings['MODEL_NAME']}",
                                settings['CNN_INPUT_DIM'],
                                confidence_thresh=settings['CONFIDENCE_THRESHOLD'])

    # Get filepaths of input images
    input_img_fnames = []
    for extension in settings['VALID_INPUT_FILETYPES']:
        input_img_fnames.extend(glob.glob(os.path.join(settings['INPUT_FOLDER'], extension)))
    input_img_fnames = sorted(list(set(input_img_fnames)))

    # Create/Clear the output directory
    if not os.path.exists(settings['OUTPUT_FOLDER']):
        os.makedirs(settings['OUTPUT_FOLDER'])
    else:
        for filename in os.listdir(settings['OUTPUT_FOLDER']):
            file_path = os.path.join(settings['OUTPUT_FOLDER'], filename)
            if os.path.isfile(file_path) and any(file_path.endswith(ext.replace('*','')) for ext in settings['VALID_INPUT_FILETYPES']):
                os.remove(file_path)
    # # Delete the error log file
    # try:
    #     os.remove('ERROR.LOG')
    # except OSError as e:
    #     pass

    # Run object detection on all the images
    count_cells(input_img_fnames, image_processor, cnn_detector)

    # Print exit message and hang until key is pressed
    run_duration = time.time()-PROGRAM_START_TIME
    print("\nFinished in %02d:%02d:%02d."%(run_duration//3600, run_duration//60, run_duration%60))
    exit_hang()
    cv2.destroyAllWindows()